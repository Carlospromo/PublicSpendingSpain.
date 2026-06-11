"""Tests del extractor IGAE mensual (sin red: URL, descubrimiento y fixture)."""

from datetime import date
from pathlib import Path
from typing import Any

import httpx
import openpyxl
import pytest

from gasto_estado.extractors import igae_mensual

DOCS = (
    "https://www.igae.pap.hacienda.gob.es/sitios/igae/es-ES/Contabilidad"
    "/ContabilidadPublica/CPE/EjecucionPresupuestaria/Documents"
)
FIXTURE = Path(__file__).parent / "fixtures" / "igae_anexo_i_muestra.xlsx"


# ---------------------------------------------------------------------------
# Composición de URL (patrón verificado 2026-06-11)
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    ("periodo", "esperado"),
    [
        ("2026-04", f"{DOCS}/MENSUAL%20ABRIL%202026%20ANEXO%20I.xlsx"),
        ("2025-12", f"{DOCS}/MENSUAL%20DICIEMBRE%202025%20ANEXO%20I.xlsx"),
        ("2019-09", f"{DOCS}/MENSUAL%20SEPTIEMBRE%202019%20ANEXO%20I.xlsx"),
    ],
)
def test_compose_anexo_i_url(periodo: str, esperado: str) -> None:
    assert igae_mensual.compose_anexo_i_url(periodo, documents_base=DOCS) == esperado


@pytest.mark.parametrize("malo", ["2026-13", "2026-00", "202604", "abril 2026", "2026-4"])
def test_periodo_invalido_falla(malo: str) -> None:
    with pytest.raises(ValueError):
        igae_mensual.parse_periodo(malo)


def test_anexo_i_filename() -> None:
    assert igae_mensual.anexo_i_filename("2026-04") == "MENSUAL ABRIL 2026 ANEXO I.xlsx"


# ---------------------------------------------------------------------------
# Descubrimiento del último periodo (red mockeada)
# ---------------------------------------------------------------------------


def test_discover_latest_retrocede_hasta_el_publicado(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """A 2026-06-11, mayo y junio dan 404 y abril 200 → resuelve 2026-04."""
    monkeypatch.setattr(
        igae_mensual,
        "_source_config",
        lambda: {"url_documents": DOCS, "url_index": "https://example.test/index.aspx"},
    )
    probed: list[str] = []

    def fake_head(url: str, **_: Any) -> httpx.Response:
        probed.append(url)
        status = 200 if "ABRIL%202026" in url else 404
        return httpx.Response(status, request=httpx.Request("HEAD", url))

    monkeypatch.setattr(igae_mensual.httpx, "head", fake_head)
    periodo = igae_mensual.discover_latest_periodo(today=date(2026, 6, 11))
    assert periodo == "2026-04"
    assert probed == [
        f"{DOCS}/MENSUAL%20JUNIO%202026%20ANEXO%20I.xlsx",
        f"{DOCS}/MENSUAL%20MAYO%202026%20ANEXO%20I.xlsx",
        f"{DOCS}/MENSUAL%20ABRIL%202026%20ANEXO%20I.xlsx",
    ]


def test_discover_latest_fallback_scraping(monkeypatch: pytest.MonkeyPatch) -> None:
    """Si ningún sondeo resuelve, se scrapea la página índice del año."""
    monkeypatch.setattr(
        igae_mensual,
        "_source_config",
        lambda: {"url_documents": DOCS, "url_index": "https://example.test/index.aspx"},
    )
    monkeypatch.setattr(
        igae_mensual.httpx,
        "head",
        lambda url, **_: httpx.Response(404, request=httpx.Request("HEAD", url)),
    )
    html = (
        '<a href="/x/MENSUAL%20ENERO%202026%20ANEXO%20I.xlsx">a</a>'
        '<a href="/x/MENSUAL%20MARZO%202026%20ANEXO%20I.xlsx">b</a>'
        '<a href="/x/MENSUAL%20DICIEMBRE%202018%20ANEXO%20I%20%28EXCEL%29.xlsx">c</a>'
    )
    monkeypatch.setattr(
        igae_mensual.httpx,
        "get",
        lambda url, **_: httpx.Response(200, text=html, request=httpx.Request("GET", url)),
    )
    assert igae_mensual.discover_latest_periodo(today=date(2026, 6, 11)) == "2026-03"


def test_extract_compone_url_y_subdir_por_periodo(monkeypatch: pytest.MonkeyPatch) -> None:
    """extract(--periodo) compone URL y guarda bajo el subdirectorio del periodo."""
    monkeypatch.setattr(
        igae_mensual,
        "_source_config",
        lambda: {"url_documents": DOCS, "url_index": "https://example.test/index.aspx"},
    )
    llamadas: list[dict[str, Any]] = []

    def fake_download(url: str, **kwargs: Any) -> Path:
        llamadas.append({"url": url, **kwargs})
        return Path("/tmp/fake.xlsx")

    monkeypatch.setattr(igae_mensual, "download_to_raw", fake_download)
    igae_mensual.extract(periodo="2026-04", raw_dir=Path("/tmp/raw"))
    assert llamadas == [
        {
            "url": f"{DOCS}/MENSUAL%20ABRIL%202026%20ANEXO%20I.xlsx",
            "fuente": "igae_mensual",
            "filename": "MENSUAL ABRIL 2026 ANEXO I.xlsx",
            "raw_dir": Path("/tmp/raw"),
            "subdir": "2026-04",
        }
    ]


# ---------------------------------------------------------------------------
# Fixture: guarda la maquetación que el parser del Prompt 5 espera
# ---------------------------------------------------------------------------


def test_fixture_anexo_i_conserva_maquetacion() -> None:
    wb = openpyxl.load_workbook(FIXTURE, read_only=True)
    assert wb.sheetnames == ["S5", "S01", "S04", "S15"]
    ws = wb["S15"]
    ws.reset_dimensions()
    filas = list(ws.iter_rows(values_only=True, max_row=3))
    cabecera = [str(c) for c in filas[1] if c]
    assert any("CRÉDITOS" in c and "INICIALES" in c for c in cabecera)
    assert any("OBLIGACIONES RECONOCIDAS NETAS" in c for c in cabecera)
    assert any(c == "APLICACIÓN PGE" for c in cabecera)
    wb.close()
