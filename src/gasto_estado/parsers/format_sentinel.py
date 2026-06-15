"""Centinela de formato pre-parser (Fase 7, resiliencia operativa).

Verifica, ANTES de parsear, que un fichero recién descargado tiene la estructura
esperada para su fuente y vintage. No parsea: solo verifica que el formato es el
esperado. Si falla, produce un diagnóstico detallado (qué esperaba vs qué
encontró) y el pipeline se detiene antes de llegar al parser, con notificación.

Esto convierte un "el parser explotó con un traceback críptico" en un "la IGAE
cambió la maquetación del Anexo I: las hojas ahora se llaman Sec01 en vez de
S01, hay que investigar un posible nuevo vintage".

Cada verificador atrapa TODAS las excepciones internamente y devuelve un
``DiagnosticoFormato`` con ``valido=False``, nunca lanza.
"""

from __future__ import annotations

import io
import json
import re
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any


@dataclass(frozen=True)
class DiagnosticoFormato:
    """Resultado del centinela: fuente, si es válido, y diagnóstico si no."""

    fuente: str
    valido: bool
    mensaje: str
    detalles: dict[str, Any] = field(default_factory=dict)


def _ok(fuente: str, **detalles: Any) -> DiagnosticoFormato:
    return DiagnosticoFormato(fuente=fuente, valido=True, mensaje="", detalles=detalles)


def _fallo(fuente: str, mensaje: str, **detalles: Any) -> DiagnosticoFormato:
    return DiagnosticoFormato(fuente=fuente, valido=False, mensaje=mensaje, detalles=detalles)


# ---------------------------------------------------------------------------
# IGAE Anexo I
# ---------------------------------------------------------------------------

_ZIP_MAGIC = b"PK\x03\x04"
_OLE2_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
_SECTION_SHEET_RE = re.compile(r"^S\d{2}$")
_CREDITO_RE = re.compile(r"CR.DITO", re.IGNORECASE)


def verificar_igae(path: Path) -> DiagnosticoFormato:
    """Verifica la estructura del Anexo I antes de parsear."""
    try:
        with path.open("rb") as f:
            magic = f.read(8)

        if magic.startswith(_ZIP_MAGIC):
            import openpyxl

            wb = openpyxl.load_workbook(path, read_only=True)
            try:
                sheets = set(wb.sheetnames)
            finally:
                wb.close()
        elif magic == _OLE2_MAGIC:
            import xlrd

            book = xlrd.open_workbook(str(path))
            sheets = set(book.sheet_names())
        else:
            return _fallo(
                "igae_mensual",
                f"Contenedor no reconocido: ni OOXML (.xlsx) ni BIFF (.xls). "
                f"Magic bytes: {magic[:4].hex()}",
                magic=magic[:4].hex(),
            )

        if "S5" not in sheets:
            candidatas = [s for s in sheets if re.match(r"^[Ss]", s)]
            return _fallo(
                "igae_mensual",
                f"Falta la hoja resumen 'S5'. "
                f"Hojas encontradas con prefijo S: {sorted(candidatas)[:10]}. "
                f"¿La IGAE renombró las hojas? Investigar posible nuevo vintage.",
                hojas=sorted(sheets),
            )

        snn_sheets = [s for s in sheets if _SECTION_SHEET_RE.match(s)]
        if not snn_sheets:
            todas = sorted(sheets)
            similares = [s for s in todas if re.match(r"^(Sec|sec|S)\d", s)]
            msg = f"Sin hojas de sección (patrón SNN). Hojas: {todas[:10]}."
            if similares:
                msg += (
                    f" Las hojas ahora se llaman {similares[:3]} en vez de S01, "
                    f"investigar posible nuevo vintage."
                )
            return _fallo("igae_mensual", msg, hojas=todas)

        # Verificar cabecera de S5: buscar texto con "CRÉDITO" en primeras filas
        if magic.startswith(_ZIP_MAGIC):
            import openpyxl

            wb = openpyxl.load_workbook(path, read_only=True)
            try:
                ws = wb["S5"]
                tiene_cabecera = False
                for row in ws.iter_rows(max_row=3, values_only=True):
                    for cell in row:
                        if cell and isinstance(cell, str) and _CREDITO_RE.search(cell):
                            tiene_cabecera = True
                            break
                    if tiene_cabecera:
                        break
            finally:
                wb.close()
            if not tiene_cabecera:
                return _fallo(
                    "igae_mensual",
                    "La hoja S5 no contiene la cabecera esperada (CRÉDITO) en las "
                    "primeras 3 filas. La maquetación puede haber cambiado.",
                    hojas=sorted(sheets),
                )

        # Verificar que detect.py reconoce el vintage
        from gasto_estado.parsers.igae.detect import detect_vintage as _detect

        try:
            vintage = _detect(path)
        except ValueError as exc:
            return _fallo(
                "igae_mensual",
                f"detect.py no reconoce el vintage: {exc}",
                hojas=sorted(sheets),
            )

        return _ok(
            "igae_mensual",
            vintage=vintage,
            hojas=sorted(sheets),
            n_secciones=len(snn_sheets),
        )

    except Exception as exc:
        return _fallo("igae_mensual", f"Error inesperado al verificar: {exc}")


# ---------------------------------------------------------------------------
# PLACSP
# ---------------------------------------------------------------------------

_XML_MAGIC = b"<?xm"
_ATOM_NS = "http://www.w3.org/2005/Atom"
_CODICE2_NS = "urn:dgpe:names:draft:codice:schema:xsd:CommonBasicComponents-2"


def verificar_placsp(path: Path) -> DiagnosticoFormato:
    """Verifica la estructura de un fichero PLACSP (ATOM o ZIP) antes de parsear."""
    try:
        if path.suffix.lower() == ".zip":
            if not zipfile.is_zipfile(path):
                return _fallo("placsp", f"{path.name} no es un ZIP válido.")
            with zipfile.ZipFile(path) as zf:
                atoms = [m for m in zf.namelist() if m.lower().endswith(".atom")]
                if not atoms:
                    return _fallo(
                        "placsp",
                        f"ZIP sin ficheros .atom: {zf.namelist()[:5]}",
                        miembros=zf.namelist()[:10],
                    )
                contenido = zf.read(atoms[0])
        else:
            contenido = path.read_bytes()

        if not contenido.startswith(_XML_MAGIC):
            return _fallo(
                "placsp",
                f"El contenido no empieza por '<?xml' (magic: {contenido[:10]!r}). "
                f"Posible HTML de error o contenido corrupto.",
            )

        from lxml import etree

        root = etree.parse(io.BytesIO(contenido)).getroot()

        if root.tag != f"{{{_ATOM_NS}}}feed":
            return _fallo(
                "placsp",
                f"Raíz no es un feed Atom (tag={root.tag!r}). "
                f"¿Cambió el formato de entrega de PLACSP?",
                root_tag=root.tag,
            )

        namespaces = set(root.nsmap.values())
        if _CODICE2_NS not in namespaces:
            return _fallo(
                "placsp",
                f"Feed Atom sin namespace CODICE 2.x. "
                f"Namespaces encontrados: {sorted(namespaces)[:6]}. "
                f"¿PLACSP migró a CODICE 3 / eForms?",
                namespaces=sorted(namespaces)[:10],
            )

        ns = {"a": _ATOM_NS, "at": "http://purl.org/atompub/tombstones/1.0"}
        entries = root.findall("a:entry", ns)
        tombstones = root.findall("at:deleted-entry", ns)
        if not entries and not tombstones:
            return _fallo(
                "placsp",
                "Feed vacío: sin atom:entry ni at:deleted-entry.",
            )

        return _ok(
            "placsp",
            n_entries=len(entries),
            n_tombstones=len(tombstones),
            namespaces=sorted(namespaces)[:6],
        )

    except Exception as exc:
        return _fallo("placsp", f"Error inesperado al verificar: {exc}")


# ---------------------------------------------------------------------------
# BDNS
# ---------------------------------------------------------------------------

_BDNS_REQUIRED_KEYS = {"id", "fechaConcesion", "beneficiario", "importe"}


def verificar_bdns(path: Path) -> DiagnosticoFormato:
    """Verifica la estructura de una página JSON de la BDNS."""
    try:
        datos = json.loads(path.read_text(encoding="utf-8"))
        if not isinstance(datos, dict):
            return _fallo("bdns", f"El JSON no es un objeto (tipo: {type(datos).__name__}).")

        if "content" not in datos:
            claves = sorted(datos.keys())[:10]
            return _fallo(
                "bdns",
                f"Falta la clave 'content' en la respuesta. "
                f"Claves encontradas: {claves}. ¿Cambió el esquema del API?",
                claves=claves,
            )

        content = datos["content"]
        if not isinstance(content, list):
            return _fallo("bdns", f"'content' no es una lista (tipo: {type(content).__name__}).")

        if not content:
            return _ok("bdns", n_items=0)

        muestra = content[0]
        if not isinstance(muestra, dict):
            return _fallo("bdns", f"Los items no son objetos (tipo: {type(muestra).__name__}).")

        claves_item = set(muestra.keys())
        ausentes = _BDNS_REQUIRED_KEYS - claves_item
        if ausentes:
            return _fallo(
                "bdns",
                f"Campos esperados ausentes en los items: {sorted(ausentes)}. "
                f"Campos presentes: {sorted(claves_item)[:15]}. "
                f"¿Cambió el esquema de concesiones?",
                claves_presentes=sorted(claves_item),
                claves_ausentes=sorted(ausentes),
            )

        return _ok("bdns", n_items=len(content), claves_item=sorted(claves_item)[:15])

    except json.JSONDecodeError as exc:
        return _fallo("bdns", f"JSON inválido: {exc}")
    except Exception as exc:
        return _fallo("bdns", f"Error inesperado al verificar: {exc}")


# ---------------------------------------------------------------------------
# BOE
# ---------------------------------------------------------------------------


def verificar_boe(path: Path) -> DiagnosticoFormato:
    """Verifica la estructura de un sumario o disposición BOE (XML)."""
    try:
        contenido = path.read_bytes()
        if not contenido.startswith(_XML_MAGIC):
            return _fallo(
                "boe_sumario",
                f"No es XML (magic: {contenido[:10]!r}). Posible HTML de error de boe.es.",
            )

        from xml.etree import ElementTree as ET

        root = ET.fromstring(contenido)

        # Sumario: response/data/sumario/diario
        sumario = root.find(".//sumario/diario")
        # Disposición: documento/metadatos
        metadatos = root.find(".//documento/metadatos")

        if sumario is None and metadatos is None:
            tag = root.tag
            hijos = [c.tag for c in root][:10]
            return _fallo(
                "boe_sumario",
                f"XML sin estructura de sumario (//sumario/diario) ni de disposición "
                f"(//documento/metadatos). Raíz: {tag}, hijos: {hijos}.",
                root_tag=tag,
                hijos=hijos,
            )

        if sumario is not None:
            secciones = sumario.findall(".//seccion")
            if not secciones:
                return _fallo(
                    "boe_sumario",
                    "Sumario sin nodos 'seccion'. ¿Cambió la estructura XML del BOE?",
                )
            tiene_departamento = any(s.find("departamento") is not None for s in secciones)
            if not tiene_departamento:
                return _fallo(
                    "boe_sumario",
                    "Las secciones no contienen 'departamento'. "
                    "La estructura del sumario puede haber cambiado.",
                )
            return _ok("boe_sumario", tipo="sumario", n_secciones=len(secciones))

        return _ok("boe_sumario", tipo="disposicion")

    except ET.ParseError as exc:
        return _fallo("boe_sumario", f"XML inválido: {exc}")
    except Exception as exc:
        return _fallo("boe_sumario", f"Error inesperado al verificar: {exc}")


# ---------------------------------------------------------------------------
# Consejo de Ministros
# ---------------------------------------------------------------------------

_SUMARIO_RE = re.compile(r"SUMARIO", re.IGNORECASE)


def verificar_consejo(path: Path) -> DiagnosticoFormato:
    """Verifica la estructura de una referencia del Consejo de Ministros (HTML)."""
    try:
        from lxml import html as lxml_html

        contenido = path.read_text(encoding="utf-8")
        doc = lxml_html.fromstring(contenido)

        from lxml.html import HtmlElement

        h2_elems: list[HtmlElement] = doc.xpath("//h2")  # type: ignore[assignment]
        sumario_h2: HtmlElement | None = None
        for h2 in h2_elems:
            texto = h2.text_content().strip()
            if _SUMARIO_RE.search(texto):
                sumario_h2 = h2
                break

        if sumario_h2 is None:
            h2_textos = [h2.text_content().strip()[:50] for h2 in h2_elems][:10]
            return _fallo(
                "consejo_ministros",
                f"La referencia no contiene el bloque SUMARIO esperado (h2). "
                f"Encabezados h2 encontrados: {h2_textos}. "
                f"¿Cambió la maquetación de La Moncloa?",
                h2_encontrados=h2_textos,
            )

        h3_elems: list[HtmlElement] = doc.xpath("//h3")  # type: ignore[assignment]
        if not h3_elems:
            return _fallo(
                "consejo_ministros",
                "Referencia con SUMARIO pero sin h3 (ministerios proponentes). "
                "La estructura interna puede haber cambiado.",
            )

        return _ok(
            "consejo_ministros",
            n_h3_ministerios=len(h3_elems),
            sumario_texto=sumario_h2.text_content().strip()[:50],
        )

    except Exception as exc:
        return _fallo("consejo_ministros", f"Error inesperado al verificar: {exc}")


# ---------------------------------------------------------------------------
# Despacho
# ---------------------------------------------------------------------------

_VERIFICADORES = {
    "igae_mensual": verificar_igae,
    "placsp": verificar_placsp,
    "bdns": verificar_bdns,
    "boe_sumario": verificar_boe,
    "consejo_ministros": verificar_consejo,
}


def verificar(fuente: str, path: Path) -> DiagnosticoFormato:
    """Verifica el formato de un fichero según su fuente."""
    fn = _VERIFICADORES.get(fuente)
    if fn is None:
        return DiagnosticoFormato(
            fuente=fuente,
            valido=True,
            mensaje=f"Sin centinela definido para '{fuente}'.",
        )
    return fn(path)
