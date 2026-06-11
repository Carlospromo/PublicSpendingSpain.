"""Parser del Anexo I de la IGAE, vintage 2021+ (hojas S01–S38).

Convierte un Anexo I real en el DataFrame canónico de detalle por aplicación
presupuestaria. SOLO raw → DataFrame (CLAUDE.md §2). Estructura verificada en
``docs/igae_anexo_i_estructura.md`` y sobre el fichero real de abril 2026.

Formato de la celda de aplicación (col ``APLICACIÓN PGE``):

    1501  923M   22501  -Estudios y trabajos técnicos
    └┬─┘  └┬─┘   └─┬─┘   └────────────┬──────────────┘
  orgánica prog.  económica         denominación
   (SSNN)        (3/5/7 díg.)

Tres detalles verificados sobre el fichero real (abril 2026) que rompen un
parser ingenuo:

- La económica tiene 3 (concepto), 5 (subconcepto) o 7 dígitos (partida).
- El separador económica↔denominación es ``  -`` (2+ espacios) en 3/5 dígitos,
  pero ``-`` pegado sin espacios en el nivel de 7 dígitos
  (``2210201-GAS NATURAL``).
- Variante TERRITORIAL: con código de provincia, orgánica+provincia+programa
  van pegados sin espacios (``150107923M`` = órg. 1501, provincia 07, prog.
  923M). Capturar ``provincia_cod`` es imprescindible: dos aplicaciones pueden
  compartir sección/servicio/programa/económica y diferir solo en provincia
  (07 vs 44), de modo que sin ella se perderían filas por clave duplicada.

Tipos de fila (ninguna se descarta sin etiquetar, CLAUDE.md §2):
- ``APLICACION``: detalle (única que va al canónico).
- ``CABECERA_PROGRAMA``: subtotal de programa (``923M-Denom``); no es detalle.
- ``SUBTOTAL_SERVICIO``: ``TOTAL SERVICIO``; se aparta para el cuadre interno.
- ``SEPARADOR``: fila en blanco entre programas.
- ``OTRO``: cabeceras de hoja u otras filas no clasificables.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet

from gasto_estado.transform.text import normalize

# Tipos de fila.
APLICACION = "APLICACION"
CABECERA_PROGRAMA = "CABECERA_PROGRAMA"
SUBTOTAL_SERVICIO = "SUBTOTAL_SERVICIO"
SEPARADOR = "SEPARADOR"
OTRO = "OTRO"

# Texto que la IGAE usa para "sin dato"; se mapea a nulo (≠ 0).
SIN_DATO = "-"

# Tolerancia del cuadre aplicaciones↔subtotal: los importes vienen con 2
# decimales (céntimos); 0,01 € por aplicación acumulado sobre cientos de filas
# justifica 1 céntimo × nº de aplicaciones. Usamos 0,5 € absolutos por servicio,
# holgado frente al redondeo y estricto frente a una aplicación perdida.
TOLERANCIA_CUADRE = 0.5

_SECTION_SHEET_RE = re.compile(r"^S\d{2}$")
# Paso 1: separa <códigos> | <económica> | <denominación>. La orgánica abre con
# 4 dígitos; la económica es el primer token de 3-7 dígitos seguido de '-'
# (con o sin espacios); la denominación es el resto (puede empezar por dígito,
# p. ej. "16ª Reposición del Fondo…"). El '.+?' perezoso garantiza que se toma
# la PRIMERA económica-antes-de-guion, que es la real.
_APLICACION_RE = re.compile(r"^(?P<codes>\d{4}.*?)\s+(?P<econ>\d{3,7})\s*-\s*(?P<denom>.+)$", re.S)
# Cabecera de programa: código de programa (3-4 alfanum.) seguido de '-'.
_CABECERA_RE = re.compile(r"^[0-9A-ZÑ]{3,4}-")

# Columnas canónicas de la fuente (CLAUDE.md §8). El Anexo I cubre solo tres
# magnitudes; comprometido y pagos llegarán de Cuadros/Anexo II (Fase posterior).
DETALLE_COLUMNS = [
    "periodo",
    "fuente",
    "fecha_captura",
    "seccion_cod",
    "servicio_cod",
    "servicio_denominacion",
    "programa_cod",
    "provincia_cod",
    "economica_cod",
    "aplicacion_denominacion",
    "credito_inicial",
    "credito_definitivo",
    "orn",
]

FUENTE = "igae_anexo_i"

# Magnitudes que esta fuente NO cubre (declaración explícita, no nulos
# silenciosos): se incorporarán desde Cuadros/Anexo II en un prompt posterior.
MAGNITUDES_NO_CUBIERTAS = ("comprometido", "pagos")


@dataclass(frozen=True)
class Aplicacion:
    seccion_cod: str
    servicio_cod: str
    programa_cod: str
    provincia_cod: str | None
    economica_cod: str
    denominacion: str


def classify_row(aplicacion_pge: object) -> str:
    """Etiqueta una fila de hoja de sección por su celda ``APLICACIÓN PGE``."""
    if aplicacion_pge is None or str(aplicacion_pge).strip() == "":
        return SEPARADOR
    text = str(aplicacion_pge)
    if text.strip().upper().startswith("TOTAL SERVICIO"):
        return SUBTOTAL_SERVICIO
    if _APLICACION_RE.match(text):
        return APLICACION
    if _CABECERA_RE.match(text):
        return CABECERA_PROGRAMA
    return OTRO


def parse_aplicacion(aplicacion_pge: str) -> Aplicacion:
    """Descompone la celda de una fila ``APLICACION`` en sus códigos canónicos.

    El bloque de códigos (orgánica + provincia opcional + programa) se desespacia
    y se segmenta por longitud, una vez separada la económica:

    - ``codigos[:4]`` = orgánica (sección 2 + servicio 2).
    - el resto, si mide ≥6, abre con provincia (2 díg.) pegada (desglose
      territorial: ``120111142A`` = órg. 1201 + prov. 11 + prog. 142A).
    - ``programa_cod`` conserva el token completo (4 ó 5 car.): el 5º carácter
      es el dígito/letra de territorialización propio de Defensa (``121M2``,
      ``121MC``); el grupo de programa estándar es ``programa_cod[:4]``.
    """
    m = _APLICACION_RE.match(aplicacion_pge)
    if m is None:  # pragma: no cover - classify_row ya lo garantiza
        raise ValueError(f"No es una aplicación válida: {aplicacion_pge!r}")
    codigos = re.sub(r"\s+", "", m.group("codes"))
    orgánica, resto = codigos[:4], codigos[4:]
    if len(resto) >= 6:
        provincia_cod: str | None = resto[:2]
        programa_cod = resto[2:]
    else:
        provincia_cod = None
        programa_cod = resto
    return Aplicacion(
        seccion_cod=orgánica[:2],
        servicio_cod=orgánica[2:],
        programa_cod=programa_cod,
        provincia_cod=provincia_cod,
        economica_cod=m.group("econ"),
        denominacion=m.group("denom").strip(),
    )


def _to_amount(value: object) -> float | None:
    """Importe IGAE → float; el texto ``-`` (sin dato) se mapea a nulo (≠ 0)."""
    if value is None:
        return None
    if isinstance(value, str):
        texto = value.strip()
        if texto in ("", SIN_DATO):
            return None
        return float(texto.replace(",", "."))
    if isinstance(value, int | float):
        return float(value)
    raise ValueError(f"Importe IGAE no numérico inesperado: {value!r}")


def _sum_amount(value: object) -> float:
    """Como ``_to_amount`` pero el nulo cuenta como 0 (para sumar y cuadrar)."""
    amount = _to_amount(value)
    return amount if amount is not None else 0.0


class CuadreError(ValueError):
    """El sumatorio de aplicaciones no cuadra con su TOTAL SERVICIO."""


def parse_anexo_i(
    path: Path,
    *,
    periodo: str,
    fecha_captura: date | None = None,
) -> pd.DataFrame:
    """Parsea un Anexo I (vintage 2021+) al DataFrame canónico de detalle.

    Recorre las hojas de sección, extrae las filas ``APLICACION`` y valida que
    cada bloque de servicio cuadra con su ``TOTAL SERVICIO`` (Paso 4); si no,
    ``CuadreError`` con el detalle del servicio descuadrado (fail loud).
    """
    fecha_captura = fecha_captura or date.today()
    # Carga completa (no read_only): el fichero declara mal sus dimensiones y en
    # modo normal openpyxl las recalcula a partir de las celdas reales.
    workbook = openpyxl.load_workbook(path, data_only=True)
    registros: list[dict[str, object]] = []
    try:
        for sheet_name in workbook.sheetnames:
            if not _SECTION_SHEET_RE.match(sheet_name):
                continue
            registros.extend(
                _parse_section(workbook[sheet_name], periodo=periodo, fecha_captura=fecha_captura)
            )
    finally:
        workbook.close()

    return pd.DataFrame(registros, columns=DETALLE_COLUMNS)


def _parse_section(ws: Worksheet, *, periodo: str, fecha_captura: date) -> list[dict[str, object]]:
    registros: list[dict[str, object]] = []
    acumulado = [0.0, 0.0, 0.0]
    servicio_denominacion: str | None = None
    n_aplicaciones = 0

    for row in ws.iter_rows(values_only=True):
        col0 = row[0]
        aplicacion_pge = row[1] if len(row) > 1 else None
        magnitudes = (
            row[2] if len(row) > 2 else None,
            row[3] if len(row) > 3 else None,
            row[4] if len(row) > 4 else None,
        )
        tipo = classify_row(aplicacion_pge)

        if tipo == APLICACION:
            servicio_denominacion = normalize(str(col0)) if col0 is not None else None
            aplicacion = parse_aplicacion(str(aplicacion_pge))
            importes = [_to_amount(m) for m in magnitudes]
            for i, importe in enumerate(importes):
                acumulado[i] += importe if importe is not None else 0.0
            n_aplicaciones += 1
            registros.append(
                {
                    "periodo": periodo,
                    "fuente": FUENTE,
                    "fecha_captura": fecha_captura,
                    "seccion_cod": aplicacion.seccion_cod,
                    "servicio_cod": aplicacion.servicio_cod,
                    "servicio_denominacion": servicio_denominacion,
                    "programa_cod": aplicacion.programa_cod,
                    "provincia_cod": aplicacion.provincia_cod,
                    "economica_cod": aplicacion.economica_cod,
                    "aplicacion_denominacion": aplicacion.denominacion,
                    "credito_inicial": importes[0],
                    "credito_definitivo": importes[1],
                    "orn": importes[2],
                }
            )
        elif tipo == SUBTOTAL_SERVICIO:
            subtotal = [_sum_amount(m) for m in magnitudes]
            _check_cuadre(ws.title, servicio_denominacion, acumulado, subtotal, n_aplicaciones)
            acumulado = [0.0, 0.0, 0.0]
            n_aplicaciones = 0

    return registros


def _check_cuadre(
    sheet: str,
    servicio: str | None,
    acumulado: list[float],
    subtotal: list[float],
    n_aplicaciones: int,
) -> None:
    diffs = [abs(acumulado[i] - subtotal[i]) for i in range(3)]
    if max(diffs) > TOLERANCIA_CUADRE:
        nombres = ("credito_inicial", "credito_definitivo", "orn")
        detalle = ", ".join(
            f"{nombres[i]}: Σaplic={acumulado[i]:.2f} vs subtotal={subtotal[i]:.2f} "
            f"(Δ={diffs[i]:.2f})"
            for i in range(3)
            if diffs[i] > TOLERANCIA_CUADRE
        )
        raise CuadreError(
            f"Descuadre en {sheet}, servicio '{servicio}' "
            f"({n_aplicaciones} aplicaciones): {detalle}"
        )
